from pathlib import Path
from selenium import webdriver
from selenium.common.exceptions import (
    ElementNotInteractableException, NoSuchElementException, StaleElementReferenceException, 
    TimeoutException, UnexpectedAlertPresentException, WebDriverException
    )
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from tkinter.filedialog import askopenfilename
from tkinter import messagebox
import time

### Get a list of projects from spreadsheet, only projects [1:] will be iterated as [0]th item is processed differently due to window handling
# def get_projects_list():
provided_excel = askopenfilename(title="Choose projects list in excel", filetypes=[("*.xlsx", "*.xlsx")])
if provided_excel == '':
    messagebox.showwarning("No file chosen", "Closing the app.")
    quit()
else:
    excel_path = Path(provided_excel)
    wb = load_workbook(excel_path)
    ws = wb.active
    column_a = ws['A']
    column_b = ws['B']
    column_c = ws['C']
    single_projects = [str(column_a[x].value) for x in range(len(column_a)) if column_a[x].value is not None]
    start_ranges = [str(column_b[x].value) for x in range(len(column_b)) if column_b[x].value is not None]
    end_ranges = [str(column_c[x].value) for x in range(len(column_c)) if column_c[x].value is not None]

    try:
        first_single_project = [single_projects[0]]
        rest_of_single_projects = single_projects[1:]
    except IndexError:
        first_single_project = []
        rest_of_single_projects = []

    try:
        first_start_range = [start_ranges[0]]
        rest_start_range = start_ranges[1:]
    except IndexError:
        first_start_range = []
        first_end_range = []

    try:
        first_end_range = [end_ranges[0]]
        rest_end_range = end_ranges[1:]
    except IndexError:
        rest_start_range = []
        rest_end_range = []


options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
serv = Service(Path.cwd()/'chromedriver.exe')
print("""
-Please do not interact with page elements in the newly opened Chrome window.\n
-If you need to stop reports submission, close the Chrome window.\n
-You may switch to other applications in the meantime.\n""")

driver = webdriver.Chrome(service=serv, options=options)
driver.get('https://...')

GENERAL_TIMEOUT = 30
INPUT_TIMEOUT = 20
wait = WebDriverWait(driver, GENERAL_TIMEOUT, poll_frequency=0.15)
parent_window = driver.current_window_handle
driver.maximize_window()

def timeout_msg():
    messagebox.showerror("Error", f"Timeout of {GENERAL_TIMEOUT}s reached. Closing the app.")

def loading_error_msg():
    messagebox.showerror("Error", "Error while loading the page.")

def input_plain(element, input):
    try:
        input_field = driver.find_element(By.XPATH,element)
        wait.until(EC.visibility_of_element_located((By.XPATH,element)))
        input_field.send_keys(input)
    except TimeoutException:
        driver.quit()
        timeout_msg()
    except (StaleElementReferenceException, NoSuchElementException, ElementNotInteractableException):
        time.sleep(2)
        input_field = driver.find_element(By.XPATH,element)
        wait.until(EC.visibility_of_element_located((By.XPATH,element)))
        input_field.send_keys(input)
    except (UnexpectedAlertPresentException, WebDriverException):
        driver.quit()
        loading_error_msg()

def input_project_from(element, input):
    try:
        input_field = driver.find_element(By.XPATH,element)
        wait.until(EC.visibility_of_element_located((By.XPATH,element)))
        input_field.clear()
        input_field.send_keys(input)
    except TimeoutException:
        driver.quit()
        timeout_msg()
    except (StaleElementReferenceException, NoSuchElementException, ElementNotInteractableException):
        time.sleep(2)
        input_field = driver.find_element(By.XPATH,element)
        wait.until(EC.visibility_of_element_located((By.XPATH,element)))
        input_field.send_keys(input)   
    except (UnexpectedAlertPresentException, WebDriverException):
        driver.quit()
        loading_error_msg()


def input_project_to(element, input):
    try:
        input_field = driver.find_element(By.XPATH,element)
        wait.until(EC.visibility_of_element_located((By.XPATH,element)))
        input_field.clear()
        timer = 0  #roughly equal to time elapsed in seconds
        while input_field.get_attribute("value") != input:
            if timer >= INPUT_TIMEOUT:
                driver.quit()
                timeout_msg()
            time.sleep(0.2)
            timer += 0.2
            try:
                input_field.send_keys(input)
            except (StaleElementReferenceException, NoSuchElementException, ElementNotInteractableException):
                time.sleep(2)
                input_field = driver.find_element(By.XPATH,element)
                wait.until(EC.element_to_be_clickable((By.XPATH,element)))
                input_field.send_keys(input)
    except (StaleElementReferenceException, NoSuchElementException, ElementNotInteractableException):
        time.sleep(2)
        input_field = driver.find_element(By.XPATH,element)
        wait.until(EC.element_to_be_clickable((By.XPATH,element)))
        input_field.send_keys(input)
    except TimeoutException:
        driver.quit()
        timeout_msg()
    except (UnexpectedAlertPresentException, WebDriverException):
        driver.quit()
        loading_error_msg()

def click(element):
    try:
        time.sleep(0.1)
        button = driver.find_element(By.XPATH,element)
        wait.until(EC.element_to_be_clickable((By.XPATH,element)))
        button.click()
    except (StaleElementReferenceException, NoSuchElementException, ElementNotInteractableException):
        time.sleep(2)
        button = driver.find_element(By.XPATH,element)
        wait.until(EC.element_to_be_clickable((By.XPATH,element)))
        button.click()
    except TimeoutException:
        driver.quit()
        timeout_msg()
    except (UnexpectedAlertPresentException, WebDriverException):
        driver.quit()
        loading_error_msg()

def back_to_parent_window():
    driver.switch_to.window(parent_window)
 
def switch_to_newest_window():
    timer = 0   #roughly equal to time elapsed in seconds
    while len(driver.window_handles) <= 1:
        if timer >= GENERAL_TIMEOUT:
            driver.quit()
            timeout_msg()
        time.sleep(1)
        timer += 1
    newest_window = driver.window_handles[-1]
    driver.switch_to.window(newest_window)

def switch_frame(element):
    try:
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.XPATH, element)))
    except TimeoutException:
        driver.quit()
        timeout_msg()


### Going directly to folder_name_search window, not passing folder's name in the main/parent window
parent_search_folders_xp = '//*[@id="FolderName__xc_1"]/a'
click(parent_search_folders_xp)
switch_to_newest_window()

### Accessing frame in the newly opened window (folder report lookup) and passing data
frame_name_xp = '//frame[@title="Content"]'
switch_frame(frame_name_xp)

folder_name_xp = '//*[@id="_LOVResFrm"]/table[1]/tbody/tr[4]/td/table/tbody/tr/td/div/div/input'
report_name = 'Cost & Revenue report (Unbilled Recon)'
report_go_button_xp = '//*[@id="_LOVResFrm"]/table[1]/tbody/tr[4]/td/table/tbody/tr/td/div/div/button'
quick_select_report_xp = '//*[@id="JegFoldersLOVVO1:Content"]/tbody/tr[2]/td[2]/a'
input_plain(folder_name_xp, report_name)
click(report_go_button_xp)
click(quick_select_report_xp)

### Back to main window
driver.switch_to.window(parent_window)
time.sleep(2)

parent_window_go_button_xp = '//*[@id="goButton"]'
wait.until(EC.element_to_be_clickable((By.XPATH,parent_window_go_button_xp)))
click(parent_window_go_button_xp)

parent_project_search_button_xp = '//*[@id="PROJECT_NUMBER_1__xc_1"]/a/img'
wait.until(EC.element_to_be_clickable((By.XPATH,parent_project_search_button_xp)))
click(parent_project_search_button_xp)


### First project is [0] element from the list; another window with a frame is opened for the first 'project_from' parent field
### Passing data into frame object inside the projects window. Opening the projects window is a one-off, won't open later
### Name of the frame is the same as in the previously opened child window (report folders). Selecting 'project from' field for parent
switch_to_newest_window()
switch_frame(frame_name_xp)

project_number_field_xp = '//*[@id="_LOVResFrm"]/table[1]/tbody/tr[4]/td/table/tbody/tr/td/div/div/input'
project_go_button_xp = '//*[@id="_LOVResFrm"]/table[1]/tbody/tr[4]/td/table/tbody/tr/td/div/div/button'
quick_select_project_xp = '//*[@id="SharedValueSetLOVVO1:Content"]/tbody/tr[2]/td[2]/a'


### 1: only single projects (column A)
if len(first_single_project) > 0 and len(first_start_range) == 0:
    input_plain(project_number_field_xp, first_single_project[0])
    click(project_go_button_xp)
    click(quick_select_project_xp)

    ### Back to main window for good. Still handling first project No., but now 'project to' field
    driver.switch_to.window(parent_window)

    parent_project_from_field_xp = '//*[@id="PROJECT_NUMBER_1"]'
    parent_project_to_field_xp = '//*[@id="PROJECT_NUMBER_2"]'
    submit_button_xp = '//*[@id="submitButton"]'

    input_project_to(parent_project_to_field_xp, first_single_project[0])
    click(submit_button_xp)

    ### Iteration for the rest of projects if > 1 PN on the list. No longer need to access child 'project number' window
    if len(rest_of_single_projects) > 0:
        for project in rest_of_single_projects:
            input_project_from(parent_project_from_field_xp, project)
            input_project_to(parent_project_to_field_xp, project)
            click(submit_button_xp)

### 2: single projects + ranges (columns A-C), starting with singles
elif len(first_single_project) > 0 and len(first_start_range) > 0:
    all_start_ranges = first_start_range + rest_start_range
    all_end_ranges = first_end_range + rest_end_range

    input_plain(project_number_field_xp, first_single_project[0])
    click(project_go_button_xp)
    click(quick_select_project_xp)

    ### Back to main window for good. Still handling first project No., but now 'project to' field
    driver.switch_to.window(parent_window)

    parent_project_from_field_xp = '//*[@id="PROJECT_NUMBER_1"]'
    parent_project_to_field_xp = '//*[@id="PROJECT_NUMBER_2"]'
    submit_button_xp = '//*[@id="submitButton"]'

    input_project_to(parent_project_to_field_xp, first_single_project[0])
    click(submit_button_xp)

    ### Iteration for the rest of projects if > 1 PN on the list. No longer need to access child 'project number' window
    if len(rest_of_single_projects) > 0:
        for project in rest_of_single_projects:
            input_project_from(parent_project_from_field_xp, project)
            input_project_to(parent_project_to_field_xp, project)
            click(submit_button_xp)

    for i in range(len(all_start_ranges)):
        input_project_from(parent_project_from_field_xp, all_start_ranges[i])
        input_project_to(parent_project_to_field_xp, all_end_ranges[i])
        click(submit_button_xp)    

### 3: only ranges (columns B-C)
elif len(first_single_project) == 0 and len(first_start_range) > 0:
    input_plain(project_number_field_xp, first_start_range[0])
    click(project_go_button_xp)
    click(quick_select_project_xp)

    ### Back to main window for good. Still handling first project No., but now 'project to' field
    driver.switch_to.window(parent_window)

    parent_project_from_field_xp = '//*[@id="PROJECT_NUMBER_1"]'
    parent_project_to_field_xp = '//*[@id="PROJECT_NUMBER_2"]'
    submit_button_xp = '//*[@id="submitButton"]'

    input_project_to(parent_project_to_field_xp, first_end_range[0])
    click(submit_button_xp)

    ### Iteration for the rest of projects if > 1 PN on the list. No longer need to access child 'project number' window
    if len(rest_start_range) > 0:
        for i in range(len(rest_start_range)):
            input_project_from(parent_project_from_field_xp, rest_start_range[i])
            input_project_to(parent_project_to_field_xp, rest_end_range[i])
            click(submit_button_xp)


driver.quit()
messagebox.showinfo("Done","Your reports will be delivered via email.")

#pyinstaller --add-binary "./chromedriver.exe;." cr_downloader.py --icon "icon.ico"
